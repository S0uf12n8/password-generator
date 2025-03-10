import random
import string
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook



def generate_passwd(numbers =12, digit=True, special=True, importance_level=3):
    characters = string.ascii_letters
    if digit:
        characters += string.digits
    if special:
        characters += string.punctuation

    password = ''.join(random.choice(characters) for _ in range(numbers))
    mark_parts = {"high": "s1", "medium": "208", "low": "02"}

    
    if importance_level == 3:
        selected_mark = mark_parts["high"]
    elif importance_level == 2:
        selected_mark = mark_parts["medium"]
    else:
        selected_mark = mark_parts["low"]

    
    insert_index = random.randint(0, len(password))
    mixed_password = password[:insert_index] + selected_mark + password[insert_index:]

    return mixed_password




def save_to_excel(username, website, password, filename="passwords.xlsx"):
    
    data = {
        "Username": [username],
        "Website": [website],
        "Password": [password],
        "Date Created": [datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
    }

    try:
        df = pd.read_excel(filename)  
    except FileNotFoundError:
        df = pd.DataFrame(columns=["Username", "Website", "Password", "Date Created"])

    new_df = pd.DataFrame(data)
    df = pd.concat([df, new_df], ignore_index=True)

    df.to_excel(filename, index=False)

    adjust_column_width(filename, df)
    print("✅ Password saved successfully!")

def adjust_column_width(filename, df):
    """
    Adjusts column width in the Excel file to fit the content.
    """
    wb = load_workbook(filename)
    ws = wb.active

    for col_idx, col_name in enumerate(df.columns, 1):
        max_length = max(df[col_name].astype(str).apply(len).max(), len(col_name)) + 2
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = max_length

    wb.save(filename)

username = input("Enter your username: ")
website = input("Enter the website URL: ")
importance = int(input("Enter password importance level (1-Low, 2-Medium, 3-High): "))


password_length = int(input("Enter the number of digits for the password: "))
generated_password = generate_passwd(password_length, True, True, importance)



save_to_excel(username, website, generated_password)


print(f"\n🔐 Your password for {website} is: {generated_password}")
 
